from typing import List
from app.schemas.attendance_schema import AttendanceReport
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


class ExcelService:

    # ── Paleta de colores ──────────────────────────────────────────────
    _COLOR_HEADER_BG   = "1F3864"   # Azul oscuro
    _COLOR_HEADER_FONT = "FFFFFF"   # Blanco
    _COLOR_ROW_ALT     = "DCE6F1"   # Azul muy claro (filas pares)
    _COLOR_ROW_BASE    = "FFFFFF"   # Blanco (filas impares)
    _COLOR_TOTAL_BG    = "2E75B6"   # Azul medio (fila totales)
    _COLOR_TOTAL_FONT  = "FFFFFF"
    _COLOR_BORDER      = "A6C2E8"   # Borde suave

    def generate_report_attendance_filter(
        self,
        data: List[AttendanceReport]
    ) -> BytesIO:

        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"

        # ── Estilos reutilizables ──────────────────────────────────────
        thin = Side(style="thin", color=self._COLOR_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_font    = Font(name="Arial", bold=True, size=11, color=self._COLOR_HEADER_FONT)
        header_fill    = PatternFill("solid", fgColor=self._COLOR_HEADER_BG)
        header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell_font      = Font(name="Arial", size=10)
        cell_align_c   = Alignment(horizontal="center", vertical="center")
        cell_align_l   = Alignment(horizontal="left",   vertical="center")

        total_font     = Font(name="Arial", bold=True, size=10, color=self._COLOR_TOTAL_FONT)
        total_fill     = PatternFill("solid", fgColor=self._COLOR_TOTAL_BG)
        total_align    = Alignment(horizontal="center", vertical="center")

        # ── Título del reporte (fila 1) ────────────────────────────────
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = "Reporte de Asistencias"
        title_cell.font  = Font(name="Arial", bold=True, size=14, color=self._COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ── Encabezados (fila 2) ───────────────────────────────────────
        columns = [
            ("Usuario",          22),
            ("Área",             20),
            ("Sede",             18),
            ("Fecha",            14),
            ("Hora Entrada",     14),
            ("Hora Salida",      14),
            ("Horas Trabajadas", 18),
        ]

        for col_idx, (label, width) in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[2].height = 28

        # ── Datos ──────────────────────────────────────────────────────
        for row_idx, r in enumerate(data, start=3):
            is_even = (row_idx % 2 == 0)
            row_fill = PatternFill("solid", fgColor=self._COLOR_ROW_ALT if is_even else self._COLOR_ROW_BASE)

            row_data = [
                (r.user_name,                                                    cell_align_l),
                (r.area_name,                                                    cell_align_l),
                (r.place_name,                                                   cell_align_l),
                (r.work_date.strftime("%Y-%m-%d"),                               cell_align_c),
                (r.entry_time.strftime("%H:%M:%S") if r.entry_time else "—",    cell_align_c),
                (r.exit_time.strftime("%H:%M:%S")  if r.exit_time  else "—",    cell_align_c),
                (r.total_hours or 0,                                             cell_align_c),
            ]

            for col_idx, (value, align) in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font      = cell_font
                cell.fill      = row_fill
                cell.alignment = align
                cell.border    = border

            ws.row_dimensions[row_idx].height = 20

        # ── Fila de totales ────────────────────────────────────────────
        if data:
            total_row = len(data) + 3
            last_data = len(data) + 2

            ws.merge_cells(f"A{total_row}:F{total_row}")
            label_cell = ws[f"A{total_row}"]
            label_cell.value     = "TOTAL HORAS"
            label_cell.font      = total_font
            label_cell.fill      = total_fill
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            label_cell.border    = border

            total_cell = ws.cell(
                row=total_row, column=7,
                value=f"=SUM(G3:G{last_data})"
            )
            total_cell.font      = total_font
            total_cell.fill      = total_fill
            total_cell.alignment = total_align
            total_cell.border    = border
            ws.row_dimensions[total_row].height = 22

        # ── Freeze panes (encabezado siempre visible) ──────────────────
        ws.freeze_panes = "A3"

        # ── Auto-filtro ────────────────────────────────────────────────
        ws.auto_filter.ref = f"A2:G{len(data) + 2}" if data else "A2:G2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output